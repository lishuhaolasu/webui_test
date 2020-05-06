import openpyxl
# from openpyxl.styles import PatternFill,colors
class ReadExcel(object):
    '''操作excel类'''
    def __init__(self,file_name:str,sheet_name:str):
        '''初始化实例
            参数 1： 文件名
            参数 2： 表格名
        '''
        self.file_name = file_name 
        self.sheet_name = sheet_name
        self.test_success = openpyxl.styles.PatternFill('solid',fgColor=openpyxl.styles.colors.GREEN)
        self.test_failed = openpyxl.styles.PatternFill('solid',fgColor=openpyxl.styles.colors.RED)
    def read_data(self) -> list:
        '''读取数据
            读取所有数据，返回数据的list。结构为：[({k1:v1,k2:v2,...}),(k1:v1,k2:v2,...),...]
        '''
        self.wb = openpyxl.load_workbook(self.file_name)    
        self.sheet = self.wb[self.sheet_name]               
        
        all_case = list(self.sheet.rows)                    
        title = [k.value for k in all_case[0]]              
        cases = []                                          
        for case in all_case[1:] :                          
            cs = [val.value for val in case]                
            dic = dict(zip(title,cs))                       
            cases.append(dic)                               
        return cases                                        
        
    def write_data(self,row:int,column:int,value:object) -> None:
        '''写入数据
            按cell写入数据。传入坐标：行和列，写入对应值
            写完后保存
        '''
        self.wb = openpyxl.load_workbook(self.file_name)    
        self.sheet = self.wb[self.sheet_name]               
        self.sheet.cell(row=row,column=column,value=value)
        if value == '通过':
            self.sheet.cell(row=row,column=column).fill = self.test_success
        elif value == '未通过':
            self.sheet.cell(row=row,column=column).fill = self.test_failed
        self.wb.save(self.file_name)                        